"""
每日 Excel 报告生成器
按 Schema v1.4 的五 Sheet 结构输出。
"""
from __future__ import annotations
import datetime as dt
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
NORMAL_FONT = Font(name="Arial", size=10)
SMALL_FONT = Font(name="Arial", size=9)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DISPOSITION_FILLS = {
    "TOP": PatternFill("solid", start_color="C6EFCE"),
    "CANDIDATE": PatternFill("solid", start_color="FFF2CC"),
    "WATCH": PatternFill("solid", start_color="DDEBF7"),
    "COOLING": PatternFill("solid", start_color="F8CBAD"),
    "DROP_BUT_DIGEST": PatternFill("solid", start_color="FCE4D6"),
    "DROP": PatternFill("solid", start_color="E7E6E6"),
}


def generate_report(scored_topics: list, output_path: str, discoveries=None):
    """scored_topics: list of dict with {topic, signals, scores}."""
    wb = Workbook()
    wb.remove(wb.active)

    _add_summary(wb, scored_topics, discoveries)
    _add_bucket(wb, "1_Top_Picks", scored_topics, lambda s: s["disposition"] == "TOP")
    _add_bucket(wb, "2_Candidates", scored_topics, lambda s: s["disposition"] == "CANDIDATE")
    _add_bucket(wb, "3_Watching", scored_topics, lambda s: s["disposition"] == "WATCH")
    _add_bucket(wb, "4_Cooling", scored_topics, lambda s: s["disposition"] == "COOLING")
    _add_hot_digest(wb, scored_topics)
    if discoveries:
        _add_discoveries(wb, discoveries)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _add_summary(wb, scored_topics, discoveries=None):
    ws = wb.create_sheet("0_Summary")
    today = dt.date.today().isoformat()

    rows = [
        [f"菲律宾预测市场每日话题报告 — {today}"],
        [""],
        ["话题总数", len(scored_topics)],
    ]
    counts = {}
    for s in scored_topics:
        d = s["scores"]["disposition"]
        counts[d] = counts.get(d, 0) + 1
    for disp in ["TOP", "CANDIDATE", "WATCH", "COOLING", "DROP_BUT_DIGEST", "DROP"]:
        rows.append([f"  └ {disp}", counts.get(disp, 0)])

    if discoveries:
        rows.append([""])
        rows.append(["话题发现"])
        rows.append([f"  Google News 聚类", discoveries["total_clusters"]])
        rows.append([f"  匹配已有话题", len(discoveries["matched"])])
        rows.append([f"  新发现候选", len(discoveries["new"])])

    rows.extend([
        [""],
        ["图例"],
        ["🟢 TOP", "total≥9 且无 veto；建议立即挂市场"],
        ["🟡 CANDIDATE", "总分 7-8；人工快速审核后可挂"],
        ["👀 WATCH", "5-6 或仅 T veto；明日复查"],
        ["❄️ COOLING", "已挂但热度跌破阈值；建议结算"],
        ["🔥 HOT DIGEST", "高热度但被 S/U veto；非市场内容，用于用户留存"],
        ["⚫ DROP", "低分；本次不处理"],
        [""],
        ["说明：完整维度解释见 01_bettability_classifier.md"],
    ])

    for row in rows:
        ws.append(row)

    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E78")
    for r in range(2, ws.max_row + 1):
        for c in range(1, 3):
            cell = ws.cell(r, c)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT_WRAP
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 70


def _add_bucket(wb, sheet_name, scored_topics, predicate):
    ws = wb.create_sheet(sheet_name)
    headers = ["排名", "话题", "市场建议问题", "类型", "结算源",
               "R", "S", "T", "U", "H", "总分", "Veto", "Heat",
               "新闻EN", "新闻TL", "Reddit帖", "独立源数",
               "Top 新闻", "Reddit 讨论"]
    ws.append(headers)

    matching = [s for s in scored_topics if predicate(s["scores"])]
    matching.sort(key=lambda s: (-s["scores"]["total"], -s["scores"]["heat_score"]))

    for rank, item in enumerate(matching, 1):
        topic = item["topic"]
        sig = item["signals"]
        scores = item["scores"]

        # 市场问题：用 LLM 生成的 或 第一个 market_template
        suggested = scores.get("suggested_market_question")
        if not suggested and topic.get("market_templates"):
            suggested = topic["market_templates"][0].get("question_template", "")

        mtype = topic["market_templates"][0]["type"] if topic.get("market_templates") else ""

        # 数据统计
        n_en = len(sig.get("gnews_en", []))
        n_tl = len(sig.get("gnews_tl", []))
        n_reddit = len(sig.get("reddit", []))
        # 独立新闻源数（去重 source 字段）
        sources = set()
        for x in sig.get("gnews_en", []):
            s = x.get("source", "").strip()
            if s:
                sources.add(s)
        for x in sig.get("gnews_tl", []):
            s = x.get("source", "").strip()
            if s:
                sources.add(s)
        n_sources = len(sources)

        top_news = "\n".join(f"• {x['title'][:70]} ({x.get('source','')})"
                             for x in (sig["gnews_en"][:3] + sig["gnews_tl"][:2]))
        top_reddit = "\n".join(f"• {x['title'][:70]}"
                               for x in sig["reddit"][:3])

        ws.append([
            rank,
            topic["topic_name"],
            suggested or "(待生成)",
            mtype,
            topic["resolution"]["primary_source"],
            scores["R"]["score"],
            scores["S"]["score"],
            scores["T"]["score"],
            scores["U"]["score"] if scores["U"]["score"] is not None else "-",
            scores["H"]["score"],
            scores["total"],
            ",".join(scores["veto_dimensions"]) if scores["veto_dimensions"] else "-",
            scores["heat_score"],
            n_en,
            n_tl,
            n_reddit,
            n_sources,
            top_news,
            top_reddit,
        ])

    _apply_sheet_style(ws, widths=[6, 20, 40, 14, 30, 4, 4, 4, 4, 4, 6, 10, 6,
                                   8, 8, 8, 8, 45, 40])


def _add_hot_digest(wb, scored_topics):
    ws = wb.create_sheet("5_Hot_Digest")
    headers = ["话题", "为什么不能挂市场", "热度分", "Top 3 新闻", "为什么仍有价值"]
    ws.append(headers)

    matching = [s for s in scored_topics if s["scores"]["disposition"] == "DROP_BUT_DIGEST"]
    matching.sort(key=lambda s: -s["scores"]["heat_score"])

    for item in matching:
        topic = item["topic"]
        sig = item["signals"]
        scores = item["scores"]
        veto_reasons = []
        for d in scores["veto_dimensions"]:
            veto_reasons.append(f"{d}={scores[d]['reason']}")
        top_news = "\n".join(f"• {x['title'][:70]}"
                             for x in sig["gnews_en"][:3])
        ws.append([
            topic["topic_name"],
            "；".join(veto_reasons),
            scores["heat_score"],
            top_news,
            "作为非市场内容放在 app 首页'今日热议'区域，导流到正式市场"
        ])

    _apply_sheet_style(ws, widths=[20, 50, 8, 45, 40])


def _add_discoveries(wb, discoveries):
    ws = wb.create_sheet("6_Discoveries")
    headers = ["序号", "Section", "聚类标题", "子文章数", "来源数", "来源列表",
               "发布时间", "匹配话题", "子文章标题"]
    ws.append(headers)

    # 先显示新发现（未匹配的），再显示已匹配的
    all_items = ([(item, True) for item in discoveries["new"]] +
                 [(item, False) for item in discoveries["matched"]])

    for idx, (item, is_new) in enumerate(all_items, 1):
        cluster = item["cluster"]
        sub_titles = "\n".join(f"• {a['title'][:60]} ({a['source']})"
                               for a in cluster["sub_articles"])
        ws.append([
            idx,
            cluster.get("section", ""),
            cluster["cluster_title"],
            len(cluster["sub_articles"]),
            cluster["source_count"],
            ", ".join(cluster["sources"]),
            cluster["published"],
            item["matched_topic"] or "🆕 NEW",
            sub_titles,
        ])

    NEW_FILL = PatternFill("solid", start_color="E2EFDA")
    MATCHED_FILL = PatternFill("solid", start_color="F2F2F2")

    for r in range(2, ws.max_row + 1):
        match_cell = ws.cell(r, 8)  # 匹配话题列（加了 Section 列后右移一位）
        fill = NEW_FILL if match_cell.value == "🆕 NEW" else MATCHED_FILL
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).fill = fill

    _apply_sheet_style(ws, widths=[6, 12, 50, 10, 10, 40, 22, 20, 55])


def _apply_sheet_style(ws, widths):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.font = SMALL_FONT if ws.max_column > 10 else NORMAL_FONT
            cell.alignment = LEFT_WRAP
            cell.border = BORDER
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
